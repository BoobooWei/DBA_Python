# -*- coding:utf8 -*-
"""
Created on:
@author: BoobooWei
Email: rgweiyaping@hotmail.com
Version: V.19.03.09.0
Description:
Help:
"""

# 3rd-part Modules
from openpyxl import load_workbook

class GetMyExcel:
    """
    读取excel数据
    """

    def __init__(self, excel):
        # read excel
        self.wb = load_workbook(excel)
        # get all sheet names
        # [u'Sheet1', u'Sheet2', u'Sheet3'}
        self.sheetnames = self.wb.sheetnames

    def get_sheet_data(self):
        lines = []
        for sheetname in self.sheetnames:
            sheet = self.wb[sheetname]
            # get row num
            row_num = sheet.max_row
            # get column num
            column_num = sheet.max_column
            # 获取标题 title = ['a','b']
            title = list(map(lambda x: x.value, sheet['1']))


            # 获取数据
            for row in range(2, row_num + 1):
                lines.append(list(map(lambda x: x.value, sheet[row])))
        return (title, lines)


if __name__ == "__main__":
    xlsx_list = ['cloud_resource_distribution_account.xlsx', 'financial_client_code.xlsx']
    for xlsx in xlsx_list:
        api = GetMyExcel(xlsx)
        title, lines = api.get_sheet_data()
        print(title)
        print(lines)